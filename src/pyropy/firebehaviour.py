"""firebehaviour.py

Defines the classes used to analyse fire behaviour with PyroPy.

"""

from dataclasses import dataclass
import warnings
from openpyxl import load_workbook
from pandas import DataFrame

if __name__ == '__main__':
    import spreadmodels as fbs
    # import helpers as fbh
else:
    from . import spreadmodels as fbs
    # from . import helpers as fbh

class Incident(object):
    """A wildfire incident.

    Attributes:
        df (Dataframe): Weather and model output data
        wrf (float): wind reduction factor (0-6)
        fuel_load (float): fine fuel load (t/ha)
        fhs_surf (float): surface fuels hazard score (1-4)
        fhs_n_surf (float): near surface fuels hazard score (1-4)
        fuel_height_ns (float): near surface fuel height (cm)

    """
    def __init__(self, weather_df: DataFrame):
        self.df = weather_df
        self.wrf = None
        self.fuel_load = None #t/ha
        self.fhs_surf = None
        self.fhs_n_surf = None 
        self.fuel_height_ns = None

    
    def get_params(self) -> dict:
        """Gets the model parameters that have been defined.

        Returns:
            dict: dictionary with model parameter names and values
        """
        params = list(self.__dict__.items())
        params = dict(params[1:]) #drop the dataframe
        params = {key: val for key, val in params.items() if val}
        return params

    def get_df(self) -> DataFrame:
        """

        Returns:
            DataFrame: the Incident Data as a pandas `Dataframe`
        """
        return self.df

    def get_models(self) -> list:
        """

        Returns:
            list: a list of the models that have been run for the Incident
        """
        models = {
            'forest_mk5': 'fros_mk5',
            'forest_vesta': 'fros_vesta',
        }

        return [
            key for key, val in models.items() 
            if val in self.df.columns.values
        ]

    def update_params(self, params: dict) -> None:
        """Update several model parameters using a dictionary.

        The dictionary keys must match the name of the parameter.

        Args:
            params (dict): a dictionary of the model parameters to be updated.
        """
        for key, val in params.items():
            setattr(self, key, val)

    def run_forest_mk5(self) -> None:
        """Runs the McArthur Mk5 Forest Fire Danger Meter model.

        Adds the results to the `Incident.df`
        """
        forest_mk5_params = {
            'wrf': self.wrf,
            'fuel_load': self.fuel_load,
        }
        if self.check_params(forest_mk5_params):
            self.df = fbs.ros_forest_mk5(self.df, self.wrf, self.fuel_load)

    def run_forest_vesta(self) -> None:
        """Runs the Project Vesta (fuel hazard scores) model.

        Adds the results to the `Incident.df`
        """
        forest_vesta_params = {
            'fhs_surf': self.fhs_surf,
            'fhs_n_surf': self.fhs_n_surf,
            'fuel_height_ns': self.fuel_height_ns,
        }


        if self.check_params(forest_vesta_params):
            self.df = fbs.ros_forest_vesta(
                self.df, 
                self.fhs_surf, 
                self.fhs_n_surf, 
                self.fuel_height_ns,
            )

    def print(self, head=False) -> None:
        """Prints the field headings and rows of the `Dataframe`

        Args:
            head (bool, optional): Print only the head (first 5 rows).
                Defaults to False.
        """
        if head: print(self.df.head())
        else: print(self.df)

    def check_params(self, params: dict) -> bool:
        """Checks to see if parameters have been defined.

        Args:
            params (dict): a dictionary with the parmeters to check.

        Returns:
            bool: `True` is values for the parameters have been defined,
                else `False`
        """
        incident_params = self.get_params()
        for key in params.keys():
            if not key in incident_params.keys():
                warnings.warn(f'{key} not set - run update params')
                return False
        return True

    def compare_fba_calc(self, fn: str) -> None:
        """Loads results from an FireBehaviourCalcs spreadsheet into the 
        `Incident.df`.

        Only loads the pages from FireBehaviourCalcs that have correspond 
        to models already in the `Incident.df` 

        Args:
            fn (str): path to the FireBehaviourCalcs spreadsheet

        Returns:
            None:
        """
        calc_models = {
            'fros_mk5': ['Forest(McArthur)', 'O'],
            'fros_vesta': ['Forest(VESTA)', 'P'],
        }
        models = self.get_models()
        # for m in models
        # if fbh.check_filepath(fn, suffix='xlsm'):
        wb = load_workbook(fn, data_only=True, keep_vba=True)
        for key, val in calc_models.items():
            if key in self.df.columns.values:
                model, column = val
                ws = wb[model]
                column = ws[column]
                ros_vals =  [cell.value for cell in column if cell.value]
                ros_vals = [val for val in ros_vals if (type(val) is float)]
                if len(ros_vals) == len(self.df):
                    self.df[f'{key}_calc'] = ros_vals
        return None #ros_vals
                    



if __name__ == '__main__':
    pass